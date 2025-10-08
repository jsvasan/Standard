from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
from datetime import datetime
from bson import ObjectId
import aiosmtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import base64


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Email configuration
GMAIL_EMAIL = os.environ.get('GMAIL_EMAIL')
GMAIL_PASSWORD = os.environ.get('GMAIL_PASSWORD')

# Email sending function
async def send_email_notification(admin_email: str, registration_data: dict, include_excel: bool = True):
    try:
        # Format registration data for email
        # Calculate age from date of birth (DD/MM/YYYY format)
        from datetime import datetime
        try:
            # Parse DD/MM/YYYY format
            dob_str = registration_data['personalInfo']['dateOfBirth']
            dob = datetime.strptime(dob_str, '%d/%m/%Y')
            today = datetime.today()
            age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        except Exception as e:
            logger.warning(f"Could not parse date of birth: {registration_data['personalInfo']['dateOfBirth']}")
            age = "N/A"
        
        email_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <div style="max-width: 800px; margin: 0 auto; padding: 20px; background: #f9f9f9;">
            <div style="background: #007AFF; color: white; padding: 20px; border-radius: 8px 8px 0 0;">
                <h2 style="margin: 0;">New Health Registration Submitted</h2>
            </div>
            
            <div style="background: white; padding: 30px; border-radius: 0 0 8px 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <h3 style="color: #007AFF; border-bottom: 2px solid #007AFF; padding-bottom: 10px;">Registrant's Personal Information</h3>
                <table style="width: 100%; border-collapse: collapse; margin-bottom: 30px;">
                    <tr style="background: #f5f5f5;">
                        <td style="padding: 12px; font-weight: bold; width: 40%;">Full Name:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['registrantName']}</td>
                    </tr>
                    <tr>
                        <td style="padding: 12px; font-weight: bold;">Apartment Number:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['registrantAptNumber']}</td>
                    </tr>
                    <tr style="background: #f5f5f5;">
                        <td style="padding: 12px; font-weight: bold;">Date of Birth:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['dateOfBirth']} (Age: {age} years)</td>
                    </tr>
                    <tr>
                        <td style="padding: 12px; font-weight: bold;">Mobile Phone:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['registrantPhone']}</td>
                    </tr>
                    <tr style="background: #f5f5f5;">
                        <td style="padding: 12px; font-weight: bold;">Blood Group:</td>
                        <td style="padding: 12px;"><strong style="color: #FF3B30; font-size: 16px;">{registration_data['personalInfo']['bloodGroup']}</strong></td>
                    </tr>
                </table>
                
                <h3 style="color: #007AFF; border-bottom: 2px solid #007AFF; padding-bottom: 10px;">Medical Information</h3>
                <table style="width: 100%; border-collapse: collapse; margin-bottom: 30px;">
                    <tr style="background: #f5f5f5;">
                        <td style="padding: 12px; font-weight: bold; width: 40%;">Insurance Policy Number:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['insurancePolicy'] or 'Not provided'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 12px; font-weight: bold;">Insurance Company / ECHS:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['insuranceCompany'] or 'Not provided'}</td>
                    </tr>
                    <tr style="background: #f5f5f5;">
                        <td style="padding: 12px; font-weight: bold;">Doctor's Name:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['doctorName'] or 'Not provided'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 12px; font-weight: bold;">Doctor's Contact:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['doctorContact'] or 'Not provided'}</td>
                    </tr>
                    <tr style="background: #f5f5f5;">
                        <td style="padding: 12px; font-weight: bold;">Hospital Name:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['hospitalName'] or 'Not provided'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 12px; font-weight: bold;">Hospital Registration Number:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['hospitalNumber'] or 'Not provided'}</td>
                    </tr>
                    <tr style="background: #f5f5f5;">
                        <td style="padding: 12px; font-weight: bold; vertical-align: top;">Current Ailments:</td>
                        <td style="padding: 12px;">{registration_data['personalInfo']['currentAilments'] or 'None reported'}</td>
                    </tr>
                </table>
                
                <h3 style="color: #007AFF; border-bottom: 2px solid #007AFF; padding-bottom: 10px; margin-top: 30px;">Buddies Information</h3>
                """
        
        for idx, buddy in enumerate(registration_data['buddies'], 1):
            bg_color = '#f5f5f5' if idx % 2 == 0 else 'white'
            email_body += f"""
                <div style="background: {bg_color}; padding: 15px; margin-bottom: 15px; border-radius: 5px; border-left: 4px solid #34C759;">
                    <h4 style="margin: 0 0 10px 0; color: #34C759;">Buddy {idx}</h4>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px; font-weight: bold; width: 35%;">Name:</td>
                            <td style="padding: 8px;">{buddy['name']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold;">Phone:</td>
                            <td style="padding: 8px;">{buddy['phone']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold;">Email:</td>
                            <td style="padding: 8px;">{buddy['email']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold;">Apartment Number:</td>
                            <td style="padding: 8px;">{buddy['aptNumber']}</td>
                        </tr>
                    </table>
                </div>
            """
        
        email_body += """
                <h3 style="color: #007AFF; border-bottom: 2px solid #007AFF; padding-bottom: 10px; margin-top: 30px;">Next of Kin Contacts</h3>
                """
        
        for idx, kin in enumerate(registration_data['nextOfKin'], 1):
            bg_color = '#f5f5f5' if idx % 2 == 0 else 'white'
            email_body += f"""
                <div style="background: {bg_color}; padding: 15px; margin-bottom: 15px; border-radius: 5px; border-left: 4px solid #FF9500;">
                    <h4 style="margin: 0 0 10px 0; color: #FF9500;">Contact {idx}</h4>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px; font-weight: bold; width: 35%;">Name:</td>
                            <td style="padding: 8px;">{kin['name']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold;">Phone:</td>
                            <td style="padding: 8px;">{kin['phone']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; font-weight: bold;">Email:</td>
                            <td style="padding: 8px;">{kin['email']}</td>
                        </tr>
                    </table>
                </div>
            """
        
        email_body += """
            </div>
            
            <div style="background: #f0f0f0; padding: 20px; text-align: center; margin-top: 20px; border-radius: 8px;">
                <p style="margin: 0; color: #666; font-size: 14px;">
                    This registration was submitted via the Health Registration App<br>
                    Registration Date: """ + datetime.now().strftime('%B %d, %Y at %I:%M %p') + """
                </p>
            </div>
        </div>
        </body>
        </html>
        """
        
        # Create email message
        message = MIMEMultipart('mixed')
        message['From'] = GMAIL_EMAIL
        message['To'] = admin_email
        message['Subject'] = f"New Buddy Registration - {registration_data['personalInfo']['registrantName']}"
        
        # Add HTML body
        html_part = MIMEText(email_body, 'html')
        message.attach(html_part)
        
        # Add Excel attachment if requested
        if include_excel:
            try:
                # Create Excel file with this registration
                excel_data = create_excel_from_registrations([registration_data])
                
                # Create attachment
                excel_attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                excel_attachment.set_payload(excel_data)
                encoders.encode_base64(excel_attachment)
                
                # Set filename and headers
                registrant_name = registration_data['personalInfo']['registrantName'].replace(' ', '_')
                current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"Registration_{registrant_name}_{current_date}.xlsx"
                
                excel_attachment.add_header(
                    'Content-Disposition',
                    f'attachment; filename="{filename}"'
                )
                
                message.attach(excel_attachment)
                logger.info(f"Excel attachment created: {filename}")
            except Exception as e:
                logger.error(f"Failed to create Excel attachment: {str(e)}")
        
        # Send email via Gmail SMTP
        await aiosmtplib.send(
            message,
            hostname='smtp.gmail.com',
            port=587,
            username=GMAIL_EMAIL,
            password=GMAIL_PASSWORD,
            start_tls=True
        )
        
        logger.info(f"Email sent successfully to {admin_email}")
        return True
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")
        return False

# Admin confirmation email function
async def send_admin_confirmation_email(admin_data: dict, plain_password: str):
    try:
        email_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; background: #f9f9f9;">
            <div style="background: #007AFF; color: white; padding: 30px; border-radius: 8px 8px 0 0; text-align: center;">
                <h1 style="margin: 0; font-size: 28px;">🎉 Welcome Admin!</h1>
            </div>
            
            <div style="background: white; padding: 40px; border-radius: 0 0 8px 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <h2 style="color: #007AFF; margin-top: 0;">Admin Registration Successful</h2>
                
                <p style="font-size: 16px; color: #666; margin: 20px 0;">
                    You have been successfully registered as the administrator for the Health Registration App. 
                    You will now receive email notifications for all new health registrations.
                </p>
                
                <div style="background: #f5f9ff; padding: 20px; border-radius: 8px; margin: 30px 0; border-left: 4px solid #007AFF;">
                    <h3 style="margin: 0 0 15px 0; color: #007AFF;">Your Admin Details:</h3>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold; color: #666;">Name:</td>
                            <td style="padding: 8px 0; color: #000;">{admin_data['name']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold; color: #666;">Phone:</td>
                            <td style="padding: 8px 0; color: #000;">{admin_data['phone']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold; color: #666;">Email:</td>
                            <td style="padding: 8px 0; color: #000;">{admin_data['email']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold; color: #666;">Password:</td>
                            <td style="padding: 8px 0; color: #000; font-family: monospace; background: #f0f0f0; padding: 10px !important; border-radius: 4px;">{plain_password}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold; color: #666;">Registration Date:</td>
                            <td style="padding: 8px 0; color: #000;">{admin_data['createdAt'].strftime('%B %d, %Y at %I:%M %p')}</td>
                        </tr>
                    </table>
                </div>
                
                <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #ffc107;">
                    <p style="margin: 0; color: #856404; font-size: 14px;">
                        <strong>⚠️ Important:</strong> Keep your password secure. You will need it to manage admin settings and delete the admin account if needed.
                    </p>
                </div>
                
                <div style="background: #fff3cd; padding: 20px; border-radius: 8px; margin: 30px 0; border-left: 4px solid #ffc107;">
                    <h3 style="margin: 0 0 10px 0; color: #856404;">📧 Email Notifications</h3>
                    <p style="margin: 0; color: #856404; font-size: 14px;">
                        You will receive detailed email notifications whenever a new health registration is submitted. 
                        Each email will include complete registrant information, buddy details, and next of kin contacts.
                    </p>
                </div>
                
                <div style="margin: 30px 0; padding: 20px; background: #f8f9fa; border-radius: 8px;">
                    <h3 style="margin: 0 0 15px 0; color: #333;">What You Can Do:</h3>
                    <ul style="margin: 0; padding-left: 20px; color: #666;">
                        <li style="margin: 10px 0;">Receive instant email notifications for all registrations</li>
                        <li style="margin: 10px 0;">View all registrations in the admin dashboard</li>
                        <li style="margin: 10px 0;">Export individual or all registrations</li>
                        <li style="margin: 10px 0;">Share registration details via WhatsApp</li>
                    </ul>
                </div>
                
                <div style="text-align: center; margin: 30px 0;">
                    <p style="font-size: 18px; color: #333; font-weight: 600;">
                        The Health Registration App is now ready to collect registrations!
                    </p>
                </div>
            </div>
            
            <div style="background: #f0f0f0; padding: 20px; text-align: center; margin-top: 20px; border-radius: 8px;">
                <p style="margin: 0; color: #666; font-size: 14px;">
                    Health Registration App - Admin Confirmation<br>
                    If you have any questions, please contact support.
                </p>
            </div>
        </div>
        </body>
        </html>
        """
        
        # Create email message
        message = MIMEMultipart('alternative')
        message['From'] = GMAIL_EMAIL
        message['To'] = admin_data['email']
        message['Subject'] = f"✅ Admin Registration Successful - {admin_data['name']}"
        
        html_part = MIMEText(email_body, 'html')
        message.attach(html_part)
        
        # Send email via Gmail SMTP
        await aiosmtplib.send(
            message,
            hostname='smtp.gmail.com',
            port=587,
            username=GMAIL_EMAIL,
            password=GMAIL_PASSWORD,
            start_tls=True
        )
        
        logger.info(f"Admin confirmation email sent successfully to {admin_data['email']}")
        return True
    except Exception as e:
        logger.error(f"Failed to send admin confirmation email: {str(e)}")
        return False

# Define Models
class Admin(BaseModel):
    name: str
    phone: str
    email: EmailStr
    password_hash: str
    additional_emails: List[str] = []
    createdAt: datetime = Field(default_factory=datetime.utcnow)

class AdminCreate(BaseModel):
    name: str
    phone: str
    email: EmailStr
    password: str

class AdminResponse(BaseModel):
    id: str
    name: str
    phone: str
    email: str
    additional_emails: List[str] = []
    createdAt: datetime

class AdminDeleteRequest(BaseModel):
    email: str
    password: str

class AdminRegistrationUpdateRequest(BaseModel):
    password: str
    personalInfo: dict
    buddies: List[dict]
    nextOfKin: List[dict]

class AdminRegistrationDeleteRequest(BaseModel):
    password: str

class PersonalInfo(BaseModel):
    registrantName: str
    registrantAptNumber: str
    dateOfBirth: str
    registrantPhone: str
    bloodGroup: str
    insurancePolicy: Optional[str] = ""
    insuranceCompany: Optional[str] = ""
    doctorName: Optional[str] = ""
    doctorContact: Optional[str] = ""
    hospitalName: Optional[str] = ""
    hospitalNumber: Optional[str] = ""
    currentAilments: Optional[str] = ""

class Buddy(BaseModel):
    name: str
    phone: str
    email: EmailStr
    aptNumber: str

class NextOfKin(BaseModel):
    name: str
    phone: str
    email: EmailStr

class Registration(BaseModel):
    personalInfo: PersonalInfo
    buddies: List[Buddy]
    nextOfKin: List[NextOfKin]
    createdAt: datetime = Field(default_factory=datetime.utcnow)

class RegistrationCreate(BaseModel):
    personalInfo: PersonalInfo
    buddies: List[Buddy]
    nextOfKin: List[NextOfKin]

class RegistrationResponse(BaseModel):
    id: str
    personalInfo: PersonalInfo
    buddies: List[Buddy]
    nextOfKin: List[NextOfKin]
    createdAt: datetime

# Excel generation functions
def calculate_age(date_of_birth_str: str) -> str:
    """Calculate age from date of birth string"""
    try:
        # Parse date in DD/MM/YYYY format
        day, month, year = date_of_birth_str.split('/')
        dob = datetime(int(year), int(month), int(day))
        today = datetime.now()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return str(age)
    except:
        return 'N/A'

def create_excel_from_registrations(registrations_list: List[dict], filename: str = "registrations.xlsx") -> bytes:
    """Create Excel file from registrations data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Buddy Registrations"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    
    # Headers
    headers = [
        'Registration Date', 'Full Name', 'Apt Number', 'Date of Birth', 'Age', 
        'Mobile Phone', 'Blood Group', 'Insurance Policy', 'Insurance Company',
        'Doctor Name', 'Doctor Contact', 'Hospital Name', 'Hospital Reg Number',
        'Current Ailments', 'Buddy 1 Name', 'Buddy 1 Phone', 'Buddy 1 Email',
        'Buddy 1 Apt', 'Buddy 2 Name', 'Buddy 2 Phone', 'Buddy 2 Email',
        'Buddy 2 Apt', 'Next of Kin 1 Name', 'Next of Kin 1 Phone', 
        'Next of Kin 1 Email', 'Next of Kin 2 Name', 'Next of Kin 2 Phone',
        'Next of Kin 2 Email', 'Next of Kin 3 Name', 'Next of Kin 3 Phone',
        'Next of Kin 3 Email'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row_idx, reg in enumerate(registrations_list, 2):
        personal = reg['personalInfo']
        buddies = reg['buddies']
        next_of_kin = reg['nextOfKin']
        
        # Personal info
        age = calculate_age(personal.get('dateOfBirth', ''))
        reg_date = reg['createdAt'].strftime('%d/%m/%Y') if isinstance(reg['createdAt'], datetime) else reg['createdAt']
        
        row_data = [
            reg_date,
            personal.get('registrantName', ''),
            personal.get('registrantAptNumber', ''),
            personal.get('dateOfBirth', ''),
            age,
            personal.get('registrantPhone', ''),
            personal.get('bloodGroup', ''),
            personal.get('insurancePolicy', ''),
            personal.get('insuranceCompany', ''),
            personal.get('doctorName', ''),
            personal.get('doctorContact', ''),
            personal.get('hospitalName', ''),
            personal.get('hospitalNumber', ''),
            personal.get('currentAilments', ''),
        ]
        
        # Buddies (up to 2)
        for i in range(2):
            if i < len(buddies):
                buddy = buddies[i]
                row_data.extend([
                    buddy.get('name', ''),
                    buddy.get('phone', ''),
                    buddy.get('email', ''),
                    buddy.get('aptNumber', ''),
                ])
            else:
                row_data.extend(['', '', '', ''])
        
        # Next of Kin (up to 3)
        for i in range(3):
            if i < len(next_of_kin):
                kin = next_of_kin[i]
                row_data.extend([
                    kin.get('name', ''),
                    kin.get('phone', ''),
                    kin.get('email', ''),
                ])
            else:
                row_data.extend(['', '', ''])
        
        # Add data to row
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.read()

# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Registration API"}

# Admin endpoints
@api_router.post("/admin/register", response_model=AdminResponse)
async def register_admin(admin: AdminCreate):
    try:
        # Check if admin already exists
        existing_admin = await db.admins.find_one({})
        if existing_admin:
            raise HTTPException(status_code=400, detail="Admin already exists. Only one admin is allowed.")
        
        # Hash the password
        plain_password = admin.password
        password_hash = bcrypt.hashpw(plain_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        admin_dict = admin.dict(exclude={'password'})
        admin_dict['password_hash'] = password_hash
        admin_dict['createdAt'] = datetime.utcnow()
        admin_dict['additional_emails'] = []
        
        result = await db.admins.insert_one(admin_dict)
        created_admin = await db.admins.find_one({"_id": result.inserted_id})
        
        # Send confirmation email to admin with password
        await send_admin_confirmation_email(created_admin, plain_password)
        
        return AdminResponse(
            id=str(created_admin['_id']),
            name=created_admin['name'],
            phone=created_admin['phone'],
            email=created_admin['email'],
            createdAt=created_admin['createdAt']
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error registering admin: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/admin", response_model=Optional[AdminResponse])
async def get_admin():
    try:
        admin = await db.admins.find_one({})
        if not admin:
            return None
        
        return AdminResponse(
            id=str(admin['_id']),
            name=admin['name'],
            phone=admin['phone'],
            email=admin['email'],
            additional_emails=admin.get('additional_emails', []),
            createdAt=admin['createdAt']
        )
    except Exception as e:
        logger.error(f"Error fetching admin: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/admin/delete")
async def delete_admin(request: AdminDeleteRequest):
    try:
        # Find admin with matching email
        admin = await db.admins.find_one({"email": request.email})
        
        if not admin:
            raise HTTPException(status_code=404, detail="Admin not found with this email")
        
        # Verify password
        if 'password_hash' in admin:
            if not bcrypt.checkpw(request.password.encode('utf-8'), admin['password_hash'].encode('utf-8')):
                raise HTTPException(status_code=401, detail="Incorrect password")
        
        # Delete the admin
        result = await db.admins.delete_one({"_id": admin['_id']})
        
        if result.deleted_count == 1:
            logger.info(f"Admin deleted: {admin['email']}")
            return {"message": "Admin deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to delete admin")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting admin: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/verify-password")
async def verify_admin_password(request: dict):
    try:
        admin = await db.admins.find_one({})
        
        if not admin:
            raise HTTPException(status_code=404, detail="Admin not found")
        
        password = request.get("password")
        if not password:
            raise HTTPException(status_code=400, detail="Password is required")
        
        # Verify password
        if 'password_hash' in admin:
            if bcrypt.checkpw(password.encode('utf-8'), admin['password_hash'].encode('utf-8')):
                return {"message": "Password verified successfully", "verified": True}
            else:
                raise HTTPException(status_code=401, detail="Incorrect password")
        else:
            raise HTTPException(status_code=400, detail="Password not set for admin")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error verifying password: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/admin/additional-emails")
async def update_additional_emails(request: dict):
    try:
        admin = await db.admins.find_one({})
        
        if not admin:
            raise HTTPException(status_code=404, detail="Admin not found")
        
        # Verify password
        password = request.get("password")
        if not password:
            raise HTTPException(status_code=400, detail="Password is required")
        
        if 'password_hash' in admin:
            if not bcrypt.checkpw(password.encode('utf-8'), admin['password_hash'].encode('utf-8')):
                raise HTTPException(status_code=401, detail="Incorrect password")
        
        additional_emails = request.get("additional_emails", [])
        
        # Validate max 2 emails
        if len(additional_emails) > 2:
            raise HTTPException(status_code=400, detail="Maximum 2 additional emails allowed")
        
        # Validate email formats
        from email.utils import parseaddr
        for email in additional_emails:
            if email and not parseaddr(email)[1]:
                raise HTTPException(status_code=400, detail=f"Invalid email format: {email}")
        
        # Update additional emails
        result = await db.admins.update_one(
            {"_id": admin['_id']},
            {"$set": {"additional_emails": additional_emails}}
        )
        
        if result.modified_count == 1 or result.matched_count == 1:
            logger.info(f"Additional emails updated for admin: {admin['email']}")
            return {"message": "Additional emails updated successfully", "additional_emails": additional_emails}
        else:
            raise HTTPException(status_code=500, detail="Failed to update additional emails")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating additional emails: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/registrations", response_model=RegistrationResponse)
async def create_registration(registration: RegistrationCreate):
    try:
        # Validate buddies count (1-2 buddies, at least 1 required)
        if len(registration.buddies) < 1 or len(registration.buddies) > 2:
            raise HTTPException(status_code=400, detail="Between 1 and 2 buddies are required (at least 1 is mandatory)")
        
        # Validate next of kin count
        if len(registration.nextOfKin) < 1 or len(registration.nextOfKin) > 3:
            raise HTTPException(status_code=400, detail="Between 1 and 3 next of kin contacts are required")
        
        # Check if registration already exists for this phone number
        existing_reg = await db.registrations.find_one({
            "personalInfo.registrantPhone": registration.personalInfo.registrantPhone
        })
        
        reg_dict = registration.dict()
        reg_dict['updatedAt'] = datetime.utcnow()
        
        is_update = False
        if existing_reg:
            # Update existing registration
            reg_dict['createdAt'] = existing_reg['createdAt']
            await db.registrations.update_one(
                {"_id": existing_reg['_id']},
                {"$set": reg_dict}
            )
            result_reg = await db.registrations.find_one({"_id": existing_reg['_id']})
            is_update = True
        else:
            # Create new registration
            reg_dict['createdAt'] = datetime.utcnow()
            result = await db.registrations.insert_one(reg_dict)
            result_reg = await db.registrations.find_one({"_id": result.inserted_id})
            is_update = False
        
        response_data = RegistrationResponse(
            id=str(result_reg['_id']),
            personalInfo=PersonalInfo(**result_reg['personalInfo']),
            buddies=[Buddy(**buddy) for buddy in result_reg['buddies']],
            nextOfKin=[NextOfKin(**kin) for kin in result_reg['nextOfKin']],
            createdAt=result_reg['createdAt']
        )
        
        # Send email notification to admin and additional emails
        admin = await db.admins.find_one({})
        if admin:
            # Send to primary admin email
            await send_email_notification(admin['email'], reg_dict)
            
            # Send to additional emails if configured
            if 'additional_emails' in admin and admin['additional_emails']:
                for email in admin['additional_emails']:
                    await send_email_notification(email, reg_dict)
        
        # Add update flag to response
        response_dict = response_data.dict()
        response_dict['is_update'] = is_update
        
        return response_dict
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating registration: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/registrations", response_model=List[RegistrationResponse])
async def get_all_registrations():
    try:
        registrations = await db.registrations.find().to_list(1000)
        return [
            RegistrationResponse(
                id=str(reg['_id']),
                personalInfo=PersonalInfo(**reg['personalInfo']),
                buddies=[Buddy(**buddy) for buddy in reg['buddies']],
                nextOfKin=[NextOfKin(**kin) for kin in reg['nextOfKin']],
                createdAt=reg['createdAt']
            )
            for reg in registrations
        ]
    except Exception as e:
        logger.error(f"Error fetching registrations: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/registrations/{registration_id}", response_model=RegistrationResponse)
async def get_registration_by_id(registration_id: str):
    try:
        if not ObjectId.is_valid(registration_id):
            raise HTTPException(status_code=400, detail="Invalid registration ID")
        
        reg = await db.registrations.find_one({"_id": ObjectId(registration_id)})
        
        if not reg:
            raise HTTPException(status_code=404, detail="Registration not found")
        
        return RegistrationResponse(
            id=str(reg['_id']),
            personalInfo=PersonalInfo(**reg['personalInfo']),
            buddies=[Buddy(**buddy) for buddy in reg['buddies']],
            nextOfKin=[NextOfKin(**kin) for kin in reg['nextOfKin']],
            createdAt=reg['createdAt']
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching registration: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/registrations/{registration_id}")
async def update_registration_admin(registration_id: str, request: AdminRegistrationUpdateRequest):
    """Admin endpoint to update a registration with password verification"""
    try:
        # Verify admin exists and password is correct
        admin = await db.admins.find_one({})
        if not admin:
            raise HTTPException(status_code=404, detail="Admin not found")
        
        if not bcrypt.checkpw(request.password.encode('utf-8'), admin['password_hash'].encode('utf-8')):
            raise HTTPException(status_code=401, detail="Invalid admin password")
        
        # Validate registration ID
        if not ObjectId.is_valid(registration_id):
            raise HTTPException(status_code=400, detail="Invalid registration ID")
        
        # Check if registration exists
        existing_reg = await db.registrations.find_one({"_id": ObjectId(registration_id)})
        if not existing_reg:
            raise HTTPException(status_code=404, detail="Registration not found")
        
        # Validate the updated data structure
        try:
            # Validate personal info
            PersonalInfo(**request.personalInfo)
            # Validate buddies (1-2 required)
            if len(request.buddies) < 1 or len(request.buddies) > 2:
                raise HTTPException(status_code=400, detail="Exactly 1-2 buddies are required")
            for buddy in request.buddies:
                Buddy(**buddy)
            # Validate next of kin (1-3 required) 
            if len(request.nextOfKin) < 1 or len(request.nextOfKin) > 3:
                raise HTTPException(status_code=400, detail="1-3 next of kin contacts are required")
            for kin in request.nextOfKin:
                NextOfKin(**kin)
        except Exception as validation_error:
            raise HTTPException(status_code=400, detail=f"Validation error: {str(validation_error)}")
        
        # Update registration
        update_data = {
            'personalInfo': request.personalInfo,
            'buddies': request.buddies,
            'nextOfKin': request.nextOfKin,
            'updatedAt': datetime.utcnow()
        }
        
        await db.registrations.update_one(
            {"_id": ObjectId(registration_id)},
            {"$set": update_data}
        )
        
        # Fetch updated registration
        updated_reg = await db.registrations.find_one({"_id": ObjectId(registration_id)})
        
        response_data = RegistrationResponse(
            id=str(updated_reg['_id']),
            personalInfo=PersonalInfo(**updated_reg['personalInfo']),
            buddies=[Buddy(**buddy) for buddy in updated_reg['buddies']],
            nextOfKin=[NextOfKin(**kin) for kin in updated_reg['nextOfKin']],
            createdAt=updated_reg['createdAt']
        )
        
        # Send email notifications for the update
        try:
            admin = await db.admins.find_one({})
            if admin:
                # Send to primary admin email
                await send_email_notification(admin['email'], updated_reg)
                
                # Send to additional emails if configured
                if 'additional_emails' in admin and admin['additional_emails']:
                    for email in admin['additional_emails']:
                        await send_email_notification(email, updated_reg)
                
                # Send to registrant's buddies
                for buddy in updated_reg['buddies']:
                    if buddy.get('email'):
                        await send_email_notification(buddy['email'], updated_reg)
                
                # Send to next of kin
                for kin in updated_reg['nextOfKin']:
                    if kin.get('email'):
                        await send_email_notification(kin['email'], updated_reg)
                
            logger.info(f"Update notification emails sent for registration {registration_id}")
        except Exception as email_error:
            logger.error(f"Failed to send update notification emails: {str(email_error)}")
            # Don't fail the whole request if email fails
        
        logger.info(f"Admin updated registration {registration_id}")
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating registration: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/registrations/{registration_id}")
async def delete_registration_admin(registration_id: str, request: AdminRegistrationDeleteRequest):
    """Admin endpoint to delete a registration with password verification"""
    try:
        # Verify admin exists and password is correct
        admin = await db.admins.find_one({})
        if not admin:
            raise HTTPException(status_code=404, detail="Admin not found")
        
        if not bcrypt.checkpw(request.password.encode('utf-8'), admin['password_hash'].encode('utf-8')):
            raise HTTPException(status_code=401, detail="Invalid admin password")
        
        # Validate registration ID
        if not ObjectId.is_valid(registration_id):
            raise HTTPException(status_code=400, detail="Invalid registration ID")
        
        # Check if registration exists
        existing_reg = await db.registrations.find_one({"_id": ObjectId(registration_id)})
        if not existing_reg:
            raise HTTPException(status_code=404, detail="Registration not found")
        
        # Delete registration
        result = await db.registrations.delete_one({"_id": ObjectId(registration_id)})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Registration not found or already deleted")
        
        logger.info(f"Admin deleted registration {registration_id}")
        return {"message": "Registration deleted successfully", "deleted_id": registration_id}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting registration: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
